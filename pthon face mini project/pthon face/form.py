import tkinter as tk
import numpy as np
import os
import openpyxl as op
from PIL import Image, ImageTk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import cv2
import pymongo
from bson import json_util

# Connect to MongoDB
client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client["face_attendance"]
collection = db["users"]

class Forget:
    def __init__(self, new):
        self.new = new
        self.new.title("FORGET PASSWORD")
        self.new.iconbitmap("E:/pthon face/ic.ico")
        
        lb1 = tk.Label(new, text="  FACE  ATTENDENCE  ", font=("Arial", 25, "underline", "bold"), fg="green").place(x=500, y=5)
        ph2 = ImageTk.PhotoImage(file="E:/pthon face/p3.jpg")
        lb3 = tk.Label(new, image=ph2, width=400, height=550).place(x=100, y=90)
        lb4 = tk.Label(new, text="' Face Attendance: Where every face tells a story of punctuality '", font=("arial", 10)).place(x=100, y=640)
        L = tk.Label(new, text="FORGET PASSWORD__", font=("arial", 15, "bold"), fg="gray").place(x=850, y=200)
        
        lb6 = tk.Label(new, text="USER_NAME :", font=("arial", 8, "bold"), width=15, height=2)
        lb6.place(x=850, y=300)
        self.en4 = tk.Entry(new, font=("arial", 9), borderwidth=3)
        self.en4.place(x=970, y=300)
        lb7 = tk.Label(new, text="OLD PASSWORD :", font=("arial", 8, "bold"), width=15, height=2)
        lb7.place(x=850, y=350)
        self.en5 = tk.Entry(new, font=("arial", 9), borderwidth=3, show="*")
        self.en5.place(x=970, y=350)
        lb8 = tk.Label(new, text="NEW PASSWORD :", font=("arial", 8, "bold"), width=15, height=2)
        lb8.place(x=850, y=400)
        self.en6 = tk.Entry(new, font=("arial", 9), borderwidth=3, show="*")
        self.en6.place(x=970, y=400)
        B = tk.Button(new, text="SUBMIT", width=10, bg="gray", command=self.update_password).place(x=950, y=450)
        
        self.new.mainloop()

    def update_password(self):
        username = self.en4.get()
        old_password = self.en5.get()
        new_password = self.en6.get()

        user = collection.find_one({"username": username, "password": old_password})
        if user:
            collection.update_one({"username": username}, {"set": {"password": new_password}})
            tk.messagebox.showinfo("Information", "Password updated successfully")
        else:
            tk.messagebox.showinfo("Information", "Incorrect username or old password")

class new_window:
    def __init__(self, new, named):
        self.new = new
        self.new.title(named)
        self.new.iconbitmap("E:/pthon face/ic.ico")
        
        lb1 = tk.Label(new, text="  FACE  ATTENDENCE  ", font=("Arial", 25, "underline", "bold"), fg="green").place(x=500, y=5)
        ph2 = ImageTk.PhotoImage(file="E:/pthon face/p5.webp")
        lb3 = tk.Label(new, image=ph2, width=400, height=550).place(x=100, y=90)
        lb4 = tk.Label(new, text="' Face Attendance: Where every face tells a story of punctuality '", font=("arial", 10)).place(x=100, y=640)
        L = tk.Label(new, text="REGISTRATION FORM FOR NEW USER", font=("arial", 15, "bold"), fg="gray").place(x=800, y=200)
        lb3 = tk.Label(new, text="FULL NAME :", font=("arial", 8, "bold"), width=10, height=2)
        lb3.place(x=850, y=250)
        self.en1 = tk.Entry(new, font=("arial", 9), borderwidth=3)
        self.en1.place(x=950, y=250)
        lb6 = tk.Label(new, text="USER_NAME :", font=("arial", 8, "bold"), width=10, height=2)
        lb6.place(x=850, y=300)
        self.en4 = tk.Entry(new, font=("arial", 9), borderwidth=3)
        self.en4.place(x=950, y=300)
        lb7 = tk.Label(new, text="PASSWORD :", font=("arial", 8, "bold"), width=10, height=2)
        lb7.place(x=850, y=350)
        self.en5 = tk.Entry(new, font=("arial", 9), borderwidth=3, show="*")
        self.en5.place(x=950, y=350)
        lb8 = tk.Label(new, text="RE-PASSWORD :", font=("arial", 8, "bold"), width=13, height=2)
        lb8.place(x=850, y=400)
        self.en6 = tk.Entry(new, font=("arial", 9), borderwidth=3, show="*")
        self.en6.place(x=950, y=400)
        B = tk.Button(new, text="SUBMIT", width=10, bg="gray", command=self.sub_pass_button).place(x=950, y=450)

        self.new.mainloop()

    def sub_pass_button(self):
        password = self.en5.get()
        re_password = self.en6.get()
        if password == re_password:
            new_user = {
                "full_name": self.en1.get(),
                "username": self.en4.get(),
                "password": password
            }
            collection.insert_one(new_user)
            tk.messagebox.showinfo("Information", "Your form has been submitted")
        else:
            tk.messagebox.showinfo("Information", "Passwords do not match")
            

class CameraApp:

    def __init__(self, window, window_title, video_source=0):
        self.window = window
        self.window.title(window_title)
        self.window.iconbitmap("E:\pthon face\ic.ico")
        lb1 = tk.Label(window, text="  FACE  ATTENDENCE  ", font=("Arial", 25, "underline", "bold"), fg="green").place(x=500, y=5)
        ph1 = ImageTk.PhotoImage(file="E:/pthon face/p1.png")
        lb3 = tk.Label(window, image=ph1, width=400, height=400).place(x=0, y=170)
        ph = ImageTk.PhotoImage(file="E:/pthon face/ex.png")
        lb2 = tk.Label(window, image=ph, width=200, height=100).place(x=300, y=160)
        lb4 = tk.Label(window, text="Attendance Sheet", font=("arial", 7, "bold")).place(x=365, y=260)
        para = """'Face attendance systems have transformed traditional methods.
                   of tracking time and attendance by introducing cutting-edge
                   facial recognition technology. This innovative approach 
                   replaces cumbersome manual processes, providing a seamless
                     and secure solution for organizations.'"""
        lb5 = tk.Label(window, text=para, font=("arial", 7, "bold"))
        lb5.place(x=50, y=570)
        self.video_source = video_source
        self.vid = cv2.VideoCapture(self.video_source)

        self.canvas = tk.Canvas(window, width=500, height=550)
        self.canvas.place(x=800, y=100)

        self.btn_snapshot = tk.Button(window, text="Capture for Attendance", command=self.snapshot).place(x=1000, y=600)
        self.update()
        self.window.mainloop()

    def snapshot(self):
        self.work_sheet()
        tk.messagebox.showinfo("information", "update")
        self.window.destroy()
    
    def work_sheet(self):
        workbook = Workbook()
        workbook = load_workbook('E:/pthon face/face_detection.xlsx')
        sheet = workbook.active
        self.name1 = e1.get()
        self.path = 'E:/pthon face/face_detection.xlsx'
        df = pd.read_excel(self.path)
        column_name = 'NAME'
        column_data = df[column_name]
        
        for i in column_data:
            if self.name1 == i:           
                sheet.cell(row=2, column=2, value="update")
                sheet.cell(row=2, column=3, value="update")
                break
            
        workbook.save('E:/pthon face/face_detection.xlsx')

    def update(self):
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        ret, frame = self.vid.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)

        if ret:
            self.photo = ImageTk.PhotoImage(image=Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)))
            self.canvas.create_image(0, 0, image=self.photo, anchor=tk.NW)

        self.window.after(10, self.update)

    def __del__(self):
        if self.vid.isOpened():
            self.vid.release()


def submit2():
    root1 = tk.Toplevel(win)
    root1.state("zoomed")
    appp = new_window(root1, "new user")

def submit():
    video_source = 0
    root = tk.Toplevel(win)
    root.state("zoomed")
    app = CameraApp(root, "FACE ATTENDANCE", video_source)

def submit3():
    root2 = tk.Toplevel(win)
    root2.state("zoomed")
    app1 = Forget(root2)


dic = {"vishesh_deep": "vishesh@23", "sonali_M": "sonali@12", "piyush_C": "piyush@12", "harshal_C": "harshal@12", "raj_khushal": "raj@2005"}

def on_click_submit():
    b = e1.get()
    b1 = e2.get()
  
    def value(dic, b1):
        for key, value in dic.items():
            if value == b1:
                return key

    result = value(dic, b1)
    if b == result and b1 == dic[b]:   
        tk.messagebox.showinfo("information", "successfully log in")
        submit()
    else:
        tk.messagebox.showinfo("information", "error")


win = tk.Tk()
win.state("zoomed")
win.iconbitmap("E:\pthon face\ic.ico")
win.title("FACE ATTENDANCE")
l = tk.Label(win, text="'Hello Student,", font=("Verdana", 15, "bold")).place(x=20, y=70)

lb = tk.Label(win, text=" FACE  ATTENDENCE ", font=("Arial", 25, "underline", "bold"), fg="green")
lb.place(x=500, y=5)

lb2 = tk.Label(win, text="Sign In", font=("arial", 20), fg="black").place(x=950, y=150)

l1 = tk.Label(win, text="USER NAME : ", font=("Arial", 10, "bold"), width=10, height=2)
l1.place(x=870, y=220)

e1 = tk.Entry(win, width=20, font=("Arial", 10), border=3)
e1.place(x=990, y=225)

l2 = tk.Label(win, text="PASSWORD : ", font=("Arial", 10, "bold"), width=10, height=2)
l2.place(x=870, y=270)

e2 = tk.Entry(win, width=20, font=("Arial", 10), show="*", borderwidth=3)
e2.place(x=990, y=275)

b1 = tk.Button(win, text="SUBMIT", font=("Arial", 10, "bold"), width=10, height=1, bg="#7d7d7d", command=on_click_submit).place(x=990, y=340)
b2 = tk.Button(win, text="new user", font=("Arial", 11, "underline"), fg="green", width=7, height=1, command=submit2).place(x=990, y=420)
b3 = tk.Button(win, text="forget password", font=("Arial", 11, "underline"), fg="green", width=15, height=1, command=submit3).place(x=1090, y=420)

vi = ImageTk.PhotoImage(file="E:/pthon face/p2.png")
label = tk.Label(win, image=vi, width=480, height=350, bg="red").place(x=60, y=140)

fi = "E:/pthon face/p3e.jpg"
file1 = ImageTk.PhotoImage(file=fi)
label1 = tk.Label(win, image=file1, width=290, height=290, bg="red").place(x=430, y=360)

l3 = tk.Label(win, text="'Your face is your password. Unlock the future with a glance.'", font=("Arial", 8)).place(x=80, y=500)

ll = tk.Label(win, text="Made by Vishesh_deep....", font=("arial", 7), fg="blue").place(x=1250, y=670)

win.mainloop()
